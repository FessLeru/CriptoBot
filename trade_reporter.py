import pandas as pd
import os
import json
from datetime import datetime
import traceback
from typing import Dict, Optional
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from config import REPORTS_DIR, TRADES_EXCEL_FILE, EXCEL_STYLES
from bot_logging import logger

class TradeReporter:
    def __init__(self, exchange):
        """
        Инициализация репортера торговли
        
        Args:
            exchange: объект для работы с биржей (ccxt)
        """
        self.exchange = exchange
        self.trades = []
        self.excel_path = os.path.join(REPORTS_DIR, TRADES_EXCEL_FILE)
        
        # Путь к файлу для хранения timestamp последнего обновления
        self.timestamp_file = os.path.join(REPORTS_DIR, "last_update_time.json")
        
        # Загружаем существующие сделки и timestamp последнего обновления
        self._load_existing_trades()
        self.last_update_time = self._load_last_update_time()
        
        logger.info(f"TradeReporter инициализирован. Последнее обновление: {datetime.fromtimestamp(self.last_update_time / 1000).strftime('%Y-%m-%d %H:%M:%S') if self.last_update_time else 'Нет данных'}")

    def _load_existing_trades(self):
        """Загрузка существующих сделок из Excel файла, если он существует"""
        if os.path.exists(self.excel_path):
            try:
                # Если файл существует, загружаем данные из него
                wb = load_workbook(self.excel_path)
                ws = wb.active
                
                # Извлекаем заголовки
                headers = [cell.value for cell in ws[1]]
                
                # Извлекаем данные начиная с второй строки
                self.trades = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    trade_dict = {headers[i]: value for i, value in enumerate(row) if i < len(headers)}
                    self.trades.append(trade_dict)
                
                logger.info(f"Загружены существующие сделки из {self.excel_path}: {len(self.trades)} сделок")
            except Exception as e:
                logger.error(f"Ошибка при загрузке существующих сделок: {e}")
                logger.error(traceback.format_exc())
                self.trades = []
        else:
            logger.info(f"Файл отчета {self.excel_path} не существует. Будет создан новый.")
            self.trades = []

    def _load_last_update_time(self):
        """Загрузка времени последнего обновления из JSON файла"""
        if os.path.exists(self.timestamp_file):
            try:
                with open(self.timestamp_file, 'r') as f:
                    data = json.load(f)
                    timestamp = data.get('last_update_time', 0)
                    logger.info(f"Загружено время последнего обновления: {datetime.fromtimestamp(timestamp / 1000).strftime('%Y-%m-%d %H:%M:%S') if timestamp else 'Нет данных'}")
                    return timestamp
            except Exception as e:
                logger.error(f"Ошибка при загрузке времени последнего обновления: {e}")
                logger.error(traceback.format_exc())
                return 0  # Если произошла ошибка, возвращаем 0, чтобы загрузить всю историю
        else:
            # Если файл не существует, возвращаем время недельной давности
            timestamp = int((datetime.now().timestamp() - 7 * 24 * 60 * 60) * 1000)  # 7 дней назад в миллисекундах
            logger.info(f"Файл с временем последнего обновления не найден, используем время 7 дней назад: {datetime.fromtimestamp(timestamp / 1000).strftime('%Y-%m-%d %H:%M:%S')}")
            return timestamp

    def _save_last_update_time(self, timestamp):
        """Сохранение времени последнего обновления в JSON файл"""
        try:
            with open(self.timestamp_file, 'w') as f:
                json.dump({'last_update_time': timestamp}, f)
            logger.info(f"Сохранено время последнего обновления: {datetime.fromtimestamp(timestamp / 1000).strftime('%Y-%m-%d %H:%M:%S')}")
        except Exception as e:
            logger.error(f"Ошибка при сохранении времени последнего обновления: {e}")
            logger.error(traceback.format_exc())

    async def fetch_new_trades(self):
        """Получение новых закрытых сделок с биржи с момента последнего обновления"""
        try:
            logger.info(f"Запрос истории сделок с {datetime.fromtimestamp(self.last_update_time / 1000).strftime('%Y-%m-%d %H:%M:%S')}")
            
            # Bitget требует символ для получения истории сделок, запросим историю для каждого популярного символа
            all_closed_trades = []
            
            # Получаем список всех символов из открытых позиций
            try:
                positions = await self.exchange.fetch_positions(params={"marginCoin": "USDT"})
                symbols = [pos['symbol'] for pos in positions if pos['symbol']]
                
                # Добавляем символы из активных ордеров
                open_orders = await self.exchange.fetch_open_orders(params={"marginCoin": "USDT"})
                for order in open_orders:
                    if order['symbol'] and order['symbol'] not in symbols:
                        symbols.append(order['symbol'])
                
                # Если нет активных позиций или ордеров, используем дефолтный список
                if not symbols:
                    # Используем символы по умолчанию для USDT-маржинальной торговли
                    symbols = ["BTC/USDT", "ETH/USDT"]
                
                logger.info(f"Запрос истории сделок для символов: {symbols}")
                
                for symbol in symbols:
                    try:
                        # Получаем историю сделок для конкретного символа
                        symbol_trades = await self.exchange.fetch_my_trades(symbol=symbol, since=self.last_update_time, params={"marginCoin": "USDT"})
                        logger.info(f"Получено {len(symbol_trades)} сделок для {symbol}")
                        all_closed_trades.extend(symbol_trades)
                    except Exception as symbol_error:
                        logger.error(f"Ошибка при получении сделок для {symbol}: {symbol_error}")
                        continue
            except Exception as pos_error:
                logger.error(f"Ошибка при получении позиций: {pos_error}")
                logger.error(traceback.format_exc())
                
                # Запасной вариант - используем Fetch Balance для получения списка валют
                try:
                    balance = await self.exchange.fetch_balance({'type': 'swap', 'marginCoin': 'USDT'})
                    currencies = [curr for curr in balance['total'].keys() if float(balance['total'][curr]) > 0]
                    
                    # Строим пары для USDT-маржинальной торговли
                    symbols = [f"{curr}/USDT" for curr in currencies if curr != 'USDT']
                    
                    # Если нет валют, используем основные
                    if not symbols:
                        symbols = ["BTC/USDT", "ETH/USDT", "SOL/USDT"]
                    
                    logger.info(f"Используем альтернативный список символов: {symbols}")
                    
                    for symbol in symbols:
                        try:
                            # Получаем историю сделок для конкретного символа
                            symbol_trades = await self.exchange.fetch_my_trades(symbol=symbol, since=self.last_update_time, params={"marginCoin": "USDT"})
                            logger.info(f"Получено {len(symbol_trades)} сделок для {symbol}")
                            all_closed_trades.extend(symbol_trades)
                        except Exception as symbol_error:
                            logger.error(f"Ошибка при получении сделок для {symbol}: {symbol_error}")
                            continue
                except Exception as balance_error:
                    logger.error(f"Ошибка при получении баланса: {balance_error}")
                    logger.error(traceback.format_exc())
                    return []
            
            closed_trades = all_closed_trades
            
            if not closed_trades:
                logger.info("Новых сделок не найдено")
                return []
            
            logger.info(f"Всего получено {len(closed_trades)} сделок")
                
            # Обновляем время последнего обновления (берем timestamp последней сделки + 1 мс)
            if closed_trades:
                latest_trade_time = max(trade['timestamp'] for trade in closed_trades)
                self.last_update_time = latest_trade_time + 1
                self._save_last_update_time(self.last_update_time)
            
            # Преобразуем сделки в удобный формат и добавляем только новые
            new_trades = []
            existing_trade_ids = {trade.get('trade_id') for trade in self.trades if trade.get('trade_id')}
            
            for trade in closed_trades:
                # Проверяем, что сделка уже не существует в нашем списке
                if trade['id'] in existing_trade_ids:
                    continue
                
                # Извлекаем необходимые данные
                trade_info = {
                    'trade_id': trade['id'],
                    'symbol': trade['symbol'],
                    'side': trade['side'],
                    'amount': trade['amount'],
                    'price': trade['price'],
                    'cost': trade['cost'],
                    'fee': trade['fee']['cost'] if trade['fee'] else 0,
                    'timestamp': datetime.fromtimestamp(trade['timestamp'] / 1000),
                }
                
                # Рассчитываем PnL более точным способом
                trade_info['pnl'] = self._calculate_trade_pnl(trade)
                
                # Добавляем сделку в список новых сделок
                new_trades.append(trade_info)
                
                # Добавляем сделку в общий список
                self.trades.append(trade_info)
            
            # Сохраняем обновленный список сделок
            if new_trades:
                logger.info(f"Добавлено {len(new_trades)} новых сделок")
                self._generate_excel_report()
            
            return new_trades
            
        except Exception as e:
            logger.error(f"Ошибка при получении новых сделок: {e}")
            logger.error(traceback.format_exc())
            return []
    
    def _calculate_trade_pnl(self, trade: Dict) -> float:
        """
        Рассчитывает PNL для сделки, используя различные источники данных.
        
        Args:
            trade: Данные о сделке от API биржи
            
        Returns:
            float: Значение PNL или 0, если не удалось рассчитать
        """
        try:
            # Извлекаем информацию из API ответа
            info = trade.get('info', {})
            
            # Список полей, которые могут содержать PNL в Bitget API
            pnl_fields = [
                'pnl',           # Основное поле PNL
                'profit',        # Прибыль
                'realizedPnl',   # Реализованная прибыль/убыток
                'totalPnl',      # Общий PNL
                'netProfit',     # Чистая прибыль
                'realPnl',       # Реальный PNL (возможная вариация)
                'tradePnl',      # PNL сделки
                'closePnl'       # PNL при закрытии
            ]
            
            # Пытаемся найти PNL в различных полях
            for field in pnl_fields:
                if field in info and info[field] is not None:
                    try:
                        pnl_value = float(info[field])
                        if pnl_value != 0:  # Если найдено ненулевое значение
                            logger.debug(f"PNL найден в поле '{field}': {pnl_value} для сделки {trade.get('id')}")
                            return pnl_value
                    except (ValueError, TypeError):
                        continue
            
            # Если PNL не найден в API, пытаемся рассчитать вручную для фьючерсов
            calculated_pnl = self._manual_pnl_calculation(trade)
            if calculated_pnl is not None:
                logger.debug(f"PNL рассчитан вручную: {calculated_pnl} для сделки {trade.get('id')}")
                return calculated_pnl
            
            # Логируем отсутствие PNL для отладки
            logger.warning(f"Не удалось найти или рассчитать PNL для сделки {trade.get('id')} ({trade.get('symbol')})")
            logger.debug(f"Доступные поля в info: {list(info.keys()) if info else 'info отсутствует'}")
            
            return 0.0
            
        except Exception as e:
            logger.error(f"Ошибка при расчете PNL для сделки {trade.get('id')}: {e}")
            return 0.0
    
    def _manual_pnl_calculation(self, trade: Dict) -> Optional[float]:
        """
        Ручной расчет PNL для фьючерсных сделок.
        
        Args:
            trade: Данные о сделке
            
        Returns:
            float: Рассчитанный PNL или None, если расчет невозможен
        """
        try:
            info = trade.get('info', {})
            
            # Для фьючерсов нужны данные о входной и выходной цене
            # Пытаемся извлечь данные о позиции
            side = trade.get('side')  # buy или sell
            amount = float(trade.get('amount', 0))
            price = float(trade.get('price', 0))
            
            # Для точного расчета PNL фьючерсов нужна информация о:
            # 1. Цене входа в позицию (может отличаться от цены текущей сделки)
            # 2. Размере позиции
            # 3. Направлении позиции (long/short)
            
            # Пытаемся найти дополнительную информацию в info
            entry_price = None
            position_side = None
            
            # Возможные поля для цены входа и направления позиции
            for price_field in ['entryPrice', 'avgPrice', 'openPrice', 'posSide']:
                if price_field in info:
                    if price_field == 'posSide':
                        position_side = info[price_field]
                    else:
                        try:
                            entry_price = float(info[price_field])
                        except (ValueError, TypeError):
                            continue
            
            # Если у нас есть достаточно данных для расчета
            if entry_price and position_side and amount > 0:
                # Расчет PNL для фьючерсов
                if position_side.lower() == 'long':
                    # Для длинной позиции: PNL = (цена_закрытия - цена_входа) * размер
                    pnl = (price - entry_price) * amount
                elif position_side.lower() == 'short':
                    # Для короткой позиции: PNL = (цена_входа - цена_закрытия) * размер
                    pnl = (entry_price - price) * amount
                else:
                    return None
                
                logger.debug(f"Ручной расчет PNL: {position_side} позиция, вход={entry_price}, выход={price}, размер={amount}, PNL={pnl}")
                return pnl
            
            return None
            
        except Exception as e:
            logger.error(f"Ошибка при ручном расчете PNL: {e}")
            return None

    def _generate_excel_report(self):
        """Создает Excel-отчет со всеми сделками и форматирует его"""
        try:
            # Создаем новую рабочую книгу
            wb = Workbook()
            ws = wb.active
            ws.title = "Торговые сделки"
            
            # Определяем стили
            header_font = Font(bold=True, size=12, color=EXCEL_STYLES['font_color'])
            header_fill = PatternFill(start_color=EXCEL_STYLES['header_color'], 
                                    end_color=EXCEL_STYLES['header_color'], fill_type="solid")
            
            profit_fill = PatternFill(start_color=EXCEL_STYLES['profit_color'], 
                                    end_color=EXCEL_STYLES['profit_color'], fill_type="solid")
            loss_fill = PatternFill(start_color=EXCEL_STYLES['loss_color'],
                                end_color=EXCEL_STYLES['loss_color'], fill_type="solid")
            
            center_alignment = Alignment(horizontal='center', vertical='center')
            border_style = EXCEL_STYLES['border_style']
            thin_border = Border(
                left=Side(style=border_style),
                right=Side(style=border_style),
                top=Side(style=border_style),
                bottom=Side(style=border_style)
            )
            
            # Задаем заголовки
            headers = [
                "ID сделки", "Символ", "Тип", "Объем", "Цена", "Стоимость", "Комиссия", 
                "Дата и время", "PNL"
            ]
            
            # Применяем заголовки
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = thin_border
            
            # Заполняем данными
            for row_idx, trade in enumerate(self.trades, 2):
                trade_data = [
                    trade.get('trade_id', ''),
                    trade.get('symbol', ''),
                    'ЛОНГ' if trade.get('side') == 'buy' else 'ШОРТ',
                    f"{trade.get('amount', 0):.6f}",
                    f"{trade.get('price', 0):.4f}",
                    f"{trade.get('cost', 0):.4f}",
                    f"{trade.get('fee', 0):.4f}",
                    trade.get('timestamp', '').strftime('%Y-%m-%d %H:%M:%S') if isinstance(trade.get('timestamp'), datetime) else '',
                    f"{trade.get('pnl', 0):.4f}" if trade.get('pnl') is not None else ''
                ]
                
                for col_idx, value in enumerate(trade_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = center_alignment
                    cell.border = thin_border
                    
                    # Подсветка для PNL
                    if col_idx == 9 and trade.get('pnl') is not None:
                        if trade.get('pnl', 0) > 0:
                            cell.fill = profit_fill
                        elif trade.get('pnl', 0) < 0:
                            cell.fill = loss_fill
            
            # Автоподбор ширины столбцов
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width
            
            # Сохраняем файл
            wb.save(self.excel_path)
            logger.info(f"Отчет успешно сохранен в {self.excel_path}")
            return self.excel_path
        except Exception as e:
            logger.error(f"Ошибка при создании Excel-отчета: {e}")
            logger.error(traceback.format_exc())
            return None

    def get_trade_statistics(self):
        """Расчет статистики по сделкам"""
        if not self.trades:
            return {
                'total_trades': 0,
                'winning_trades': 0,
                'losing_trades': 0,
                'win_rate': 0,
                'average_profit': 0,
                'average_loss': 0,
                'profit_factor': 0,
                'total_pnl': 0
            }

        # Создаем DataFrame из списка сделок
        df = pd.DataFrame(self.trades)
        
        # Если нет колонки pnl, возвращаем базовую статистику
        if 'pnl' not in df.columns:
            return {
                'total_trades': len(df),
                'total_volume': df['amount'].sum() if 'amount' in df.columns else 0,
                'average_price': df['price'].mean() if 'price' in df.columns else 0
            }
        
        # Расчет основной статистики
        total_trades = len(df)
        winning_trades = len(df[df['pnl'] > 0])
        losing_trades = len(df[df['pnl'] < 0])
        
        # Расчет процента выигрышных сделок
        win_rate = (winning_trades / total_trades * 100) if total_trades > 0 else 0
        
        # Расчет среднего профита и убытка
        avg_profit = df[df['pnl'] > 0]['pnl'].mean() if winning_trades > 0 else 0
        avg_loss = df[df['pnl'] < 0]['pnl'].mean() if losing_trades > 0 else 0
        
        # Расчет фактора прибыли
        total_profit = df[df['pnl'] > 0]['pnl'].sum() if winning_trades > 0 else 0
        total_loss = abs(df[df['pnl'] < 0]['pnl'].sum()) if losing_trades > 0 else 0
        profit_factor = total_profit / total_loss if total_loss > 0 else float('inf')
        
        # Общий PnL
        total_pnl = df['pnl'].sum()

        return {
            'total_trades': total_trades,
            'winning_trades': winning_trades,
            'losing_trades': losing_trades,
            'win_rate': win_rate,
            'average_profit': avg_profit,
            'average_loss': avg_loss,
            'profit_factor': profit_factor,
            'total_pnl': total_pnl
        }
    
    async def generate_trade_report(self):
        """
        Публичный метод для генерации отчета о сделках.
        Сначала получает обновленные данные, затем создает отчет.
        
        Returns:
            str: Путь к сгенерированному Excel-файлу или None в случае ошибки
        """
        try:
            logger.info("Запуск генерации отчета о сделках")
            
            # Создаем директорию для отчетов, если ее нет
            if not os.path.exists(REPORTS_DIR):
                os.makedirs(REPORTS_DIR)
                logger.info(f"Создана директория для отчетов: {REPORTS_DIR}")
                
            # Получаем новые сделки с биржи
            new_trades = await self.fetch_new_trades()
            logger.info(f"Получено {len(new_trades)} новых сделок для отчета")
            
            # Генерируем отчет вне зависимости от наличия новых сделок
            report_path = self._generate_excel_report()
            
            if report_path:
                logger.info(f"Отчет о сделках успешно создан: {report_path}")
            else:
                logger.warning("Не удалось создать отчет о сделках")
                
            return report_path
            
        except Exception as e:
            logger.error(f"Ошибка при генерации отчета о сделках: {e}")
            logger.error(traceback.format_exc())
            return None 