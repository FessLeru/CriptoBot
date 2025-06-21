"""
Модуль для работы с Telegram ботом.
"""
import os
import asyncio
from typing import Dict, List, Optional, Any, Callable
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json

from aiogram import Bot, Dispatcher, types
from aiogram.filters import Command
from aiogram.types import Message, FSInputFile
from aiogram.utils.markdown import hbold, hitalic, hcode, hpre

from bot_logging import logger
from trading.trader import Trader
from strategies.scanner import StrategyScanner
from utils.time_utils import get_all_supported_timeframes
from config import REPORTS_DIR, TRADES_EXCEL_FILE, EXCEL_STYLES
from trade_reporter import TradeReporter


class TelegramBot:
    """Класс для управления торговым ботом через Telegram."""
    
    def __init__(self, trader: Trader, scanner: StrategyScanner):
        """
        Инициализирует Telegram бота.
        
        Args:
            trader: Объект трейдера
            scanner: Объект сканера стратегий
        """
        self.token = os.getenv("TELEGRAM_BOT_TOKEN")
        if not self.token:
            raise ValueError("TELEGRAM_BOT_TOKEN отсутствует в .env файле")
            
        self.bot = Bot(token=self.token)
        self.dp = Dispatcher()
        self.trader = trader
        self.scanner = scanner
        self.target_chat_id = int(os.getenv("TARGET_CHAT_ID", 0))
        
        # Лок для предотвращения гонки условий при обработке сообщений
        self._message_lock = asyncio.Lock()
        
        # Регистрируем обработчики команд
        self._register_handlers()
        
        # Регистрируем обработчик сигналов
        self.scanner.register_signal_callback(self._handle_signal)
        
        logger.info(f"Инициализирован Telegram бот. Целевой чат: {self.target_chat_id}")
    
    def _register_handlers(self) -> None:
        """Регистрирует обработчики команд."""
        # Основные команды
        self.dp.message.register(self._cmd_start, Command("start"))
        self.dp.message.register(self._cmd_balance, Command("balance"))
        self.dp.message.register(self._cmd_stop, Command("stop"))
        self.dp.message.register(self._cmd_orders, Command("orders"))
        self.dp.message.register(self._cmd_leverage, Command("leverage"))
        self.dp.message.register(self._cmd_set_chat, Command("set_chat"))
        self.dp.message.register(self._cmd_get_id, Command("get_id"))
        self.dp.message.register(self._cmd_strategies, Command("strategies"))
        
        # Новые команды
        self.dp.message.register(self._cmd_timeframe, Command("timeframe"))
        self.dp.message.register(self._cmd_scan, Command("scan"))
        self.dp.message.register(self._cmd_report, Command("report"))
        self.dp.message.register(self._cmd_reload_data, Command("reload_data"))
        self.dp.message.register(self._cmd_check_indicators, Command("check_indicators"))
        
        logger.info("Зарегистрированы обработчики команд")
    
    async def start(self) -> None:
        """Запускает бота."""
        logger.info("Запуск Telegram бота")
        await self.dp.start_polling(self.bot)
    
    async def stop(self) -> None:
        """Останавливает бота."""
        logger.info("Остановка Telegram бота")
        await self.bot.session.close()
    
    async def _handle_signal(self, signal: Dict) -> None:
        """
        Обрабатывает сигнал от сканера стратегий.
        
        Args:
            signal: Словарь с информацией о сигнале
        """
        if not self.target_chat_id:
            logger.warning("Целевой чат не установлен. Сигнал не будет отправлен.")
            return
            
        try:
            # Формируем сообщение о сигнале
            symbol = signal["symbol"]
            signal_type = "🟢 ЛОНГ" if signal["side"] == "buy" else "🔴 ШОРТ"
            price = signal.get("price", 0)  # Получаем цену из сигнала или используем 0
            stop_loss = signal.get("stop_loss", 0)
            trail_points = signal.get("trail_points", 0)
            trail_offset = signal.get("trail_offset", 0)
            trail_mode = signal.get("trail_mode", True)
            strategy_name = signal.get("strategy_name", "Unknown")
            timeframe = signal.get("timeframe", "Unknown")
            
            # Получаем текущую цену для более точной информации
            try:
                ticker_data = await self.trader.exchange.get_ticker_price(symbol)
                current_price = ticker_data['mark']  # Используем mark price
            except Exception as e:
                logger.error(f"Не удалось получить текущую цену: {e}")
                current_price = price  # Используем цену из сигнала как запасной вариант
            
            # Формируем сообщение о сигнале
            signal_message = f"""🚨 СИГНАЛ от {strategy_name}: {signal_type} на {symbol}!
💰 Цена: {current_price:.4f} USDT
🛑 Стоп-лосс: {stop_loss:.4f} USDT
🔄 Трейлинг-стоп: {"✅ Активирован" if trail_mode else "❌ Отключен"}
⏱ Время: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
🔍 Таймфрейм: {timeframe}
"""
            
            # Отправляем сообщение о сигнале
            await self.bot.send_message(self.target_chat_id, signal_message)
            
            # Открываем сделку
            trade_result = await self.trader.open_trade(signal)
            
            # Проверяем, есть ли ошибка в результате
            if trade_result.startswith("⚠️"):
                # Если есть ошибка, просто отправляем результат
                await self.bot.send_message(self.target_chat_id, trade_result)
                return
            
            # Получаем данные по открытой позиции
            positions = await self.trader.get_active_positions()
            position_info = next((p for p in positions if p['symbol'] == symbol), None)
            
            # Информация о трейлинге для сообщения
            trail_info = ""
            if trail_mode:
                # Получаем данные о трейлинге из активных трейдов
                active_trade = self.trader.active_trades.get(symbol, {})
                trail_activation = active_trade.get('trail_activation', 0)
                trail_callback = active_trade.get('trail_callback', 0)
                
                if trail_activation and trail_callback:
                    trail_info = f"""🔄 Трейлинг-стоп:
   📈 Активация при цене: {trail_activation:.4f} USDT
   📉 Отступ: {trail_callback*100:.2f}%"""
                else:
                    trail_info = "🔄 Трейлинг-стоп: Настроен, детали недоступны"
            else:
                trail_info = "🔄 Трейлинг-стоп: Отключен"
            
            # Определение лонг или шорт
            position_side = "ЛОНГ" if signal["side"] == "buy" else "ШОРТ"
            
            # Формируем расширенное сообщение о результате открытия сделки
            trade_message = f"""✅ СДЕЛКА ОТКРЫТА на {symbol}

📊 Тип: {position_side}
💰 Цена входа: {current_price:.4f} USDT
📏 Объем: {signal.get('amount', position_info['contracts'] if position_info else 'N/A')}
🛑 Стоп-лосс: {stop_loss:.4f} USDT
{trail_info}
⏱ Время: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

🔍 Стратегия: {strategy_name} ({timeframe})
"""
            
            # Отправляем сообщение о результате
            await self.bot.send_message(self.target_chat_id, trade_message)
            
        except Exception as e:
            error_message = f"❌ Ошибка при обработке сигнала: {str(e)}"
            logger.error(error_message)
            await self.bot.send_message(self.target_chat_id, error_message)
    
    # --- Обработчики команд ---
    
    async def _cmd_start(self, message: Message) -> None:
        """Обработчик команды /start"""
        welcome_text = """
🚀 Привет! Я — торговый бот для работы с биржей Bitget!

Команды:
- 💰 Показать текущий баланс (/balance)
- 🛑 Остановить все активные сделки и 
     отменить лимитные заявки (/stop)
- 🔥 Поменять плечо (/leverage <число>)
- ✅ Посмотреть открытые позиции и 
     лимитные заявки (/orders)
- 🆔 Узнать ID чата (/get_id)
- 🔄 Установить этот чат как целевой (/set_chat)
- 📊 Список доступных стратегий (/strategies)
- ⏱ Изменить таймфрейм (/timeframe <символ> <таймфрейм>)
- 🔍 Запустить ручное сканирование (/scan <символ>)
- 📝 Получить отчет о торговле (/report)
- 📊 Перезагрузить исторические данные (/reload_data [таймфрейм] [лимит])
- 📈 Проверить индикаторы (/check_indicators)
"""
        await message.answer(welcome_text)
    
    async def _cmd_balance(self, message: Message) -> None:
        """Обработчик команды /balance"""
        try:
            async with self._message_lock:
                balance = await self.trader.exchange.get_usdt_balance()
                await message.reply(f"💰 Баланс фьючерсов: {balance:.2f} USDT")
                logger.info(f"Пользователь {message.from_user.id} запросил баланс: {balance:.2f} USDT")
        except Exception as e:
            logger.error(f"Ошибка при получении баланса: {str(e)}")
            await message.reply(f"⚠️ Ошибка при получении баланса: {str(e)}")
    
    async def _cmd_stop(self, message: Message) -> None:
        """Обработчик команды /stop"""
        try:
            async with self._message_lock:
                result = await self.trader.close_all_trades()
                
                canceled_orders = result.get("closed_orders", 0)
                closed_positions = result.get("closed_positions", 0)
                error = result.get("error", None)
                
                if error:
                    await message.reply(f"⚠️ Ошибка при закрытии сделок: {error}", parse_mode="Markdown")
                else:
                    result_message = (
                        f"✅ *Результат выполнения команды /stop:*\n\n"
                        f"📋 *Отменено ордеров:* {canceled_orders}\n"
                        f"📊 *Закрыто позиций:* {closed_positions}\n\n"
                        f"Все активные ордера отменены, позиции закрыты.\n"
                        f"Мониторинг рынка продолжает работать."
                    )
                    await message.reply(result_message, parse_mode="Markdown")
        except Exception as e:
            logger.error(f"Ошибка при выполнении команды /stop: {str(e)}")
            await message.reply(f"⚠️ Ошибка при выполнении команды /stop: {str(e)}", parse_mode="Markdown")
    
    async def _cmd_orders(self, message: Message) -> None:
        """Обработчик команды /orders"""
        try:
            async with self._message_lock:
                # Получаем активные ордера
                open_orders = await self.trader.get_open_orders()
        
                # Формируем сообщение с информацией об ордерах
                orders_info = "📋 *Активные ордера:*\n\n"
                if open_orders:
                    for order in open_orders:
                        orders_info += (
                            f"🔹 *Символ:* {order['symbol']}-{order['side']}\n"
                            f"   *Направление:* {'🟢 Лонг' if order['side'] == 'buy' else '🔴 Шорт'}\n"
                            f"   *Тип ордера:* {order['type'].capitalize()}\n"
                            f"   *Цена:* {order['price']:.4f}\n"
                            f"   *Объем:* {order['amount']:.4f}\n"
                            f"   *Исполнено:* {order['filled']:.4f}\n"
                            f"   *Осталось:* {order['remaining']:.4f}\n\n"
                        )
                else:
                    orders_info += "❌ *Активных ордеров нет.*\n\n"
        
                # Получаем открытые позиции
                positions = await self.trader.get_active_positions()
                positions_info = "📊 *Открытые позиции:*\n\n"
                if positions:
                    for position in positions:
                        if float(position['contracts']) > 0:
                            symbol = position['symbol']
                            side = '🟢 Лонг' if position['side'] == 'long' else '🔴 Шорт'
                            pnl = float(position['unrealizedPnl'])
                            pnl_emoji = "📈" if pnl >= 0 else "📉"
        
                            positions_info += (
                                f"🔹 *Символ:* {symbol}\n"
                                f"   *Направление:* {side}\n"
                                f"   *Объем:* {position['contracts']:.4f}\n"
                                f"   *Цена входа:* {position['entryPrice']:.4f}\n"
                                f"   *Текущая цена:* {position['markPrice']:.4f}\n"
                                f"   *PNL:* {pnl_emoji} {pnl:.4f} USDT\n\n"
                            )
                else:
                    positions_info += "❌ *Открытых позиций нет.*\n\n"
        
                # Объединяем информацию об ордерах и позициях
                final_message = orders_info + positions_info
        
                # Отправляем сообщение пользователю
                await message.reply(final_message, parse_mode="Markdown")
        except Exception as e:
            logger.error(f"Ошибка при получении списка ордеров: {str(e)}")
            await message.reply(f"⚠️ Ошибка: {str(e)}")
    
    async def _cmd_leverage(self, message: Message) -> None:
        """Обработчик команды /leverage"""
        try:
            text = message.text
            args = text.split()
            
            if len(args) != 2:
                await message.answer("⚠️ Используйте команду в формате: /leverage <число>")
                return
                
            try:
                new_leverage = int(args[1])
                
                if 1 <= new_leverage <= 100:
                    success = await self.trader.set_leverage(new_leverage)
                    if success:
                        await message.answer(f"✅ Плечо успешно изменено на {new_leverage}")
                    else:
                        await message.answer("⚠️ Не удалось изменить плечо")
                else:
                    await message.answer("⚠️ Плечо должно быть в диапазоне от 1 до 100.")
            except ValueError:
                await message.answer("⚠️ Пожалуйста, укажите корректное число после команды /leverage.")
        except Exception as e:
            logger.error(f"Ошибка при изменении плеча: {str(e)}")
            await message.reply(f"⚠️ Ошибка: {str(e)}")
    
    async def _cmd_set_chat(self, message: Message) -> None:
        """Обработчик команды /set_chat"""
        self.target_chat_id = message.chat.id
        await message.answer(f"✅ Этот чат установлен как получатель. ID: {self.target_chat_id}")
        logger.info(f"Установлен новый целевой чат: {self.target_chat_id}")
    
    async def _cmd_get_id(self, message: Message) -> None:
        """Обработчик команды /get_id"""
        chat_id = message.chat.id
        await message.answer(f"ID этого чата: <code>{chat_id}</code>", parse_mode="HTML")
    
    async def _cmd_strategies(self, message: Message) -> None:
        """Обработчик команды /strategies"""
        strategies_info = self.scanner.get_strategies_info()
        
        if not strategies_info:
            await message.reply("📊 *Стратегии не найдены*", parse_mode="Markdown")
            return
            
        response = "📈 *Доступные стратегии:*\n\n"
        
        for info in strategies_info:
            active_status = "✅ Активна" if info["active"] else "❌ Неактивна"
            response += (
                f"🔸 *Символ:* {info['symbol']}\n"
                f"   *Стратегия:* {info['name']}\n"
                f"   *Таймфрейм:* {info['timeframe']}\n"
                f"   *Статус:* {active_status}\n\n"
            )
            
        await message.reply(response, parse_mode="Markdown")
    
    async def _cmd_timeframe(self, message: Message) -> None:
        """Обработчик команды /timeframe"""
        text = message.text
        args = text.split()
        
        if len(args) != 3:
            # Выводим список доступных таймфреймов
            timeframes = get_all_supported_timeframes()
            tf_str = "\n".join([f"  • {tf} - {desc}" for tf, desc in timeframes.items()])
            
            await message.answer(
                f"⚠️ Используйте команду в формате: /timeframe <символ> <таймфрейм>\n\n"
                f"Доступные таймфреймы:\n{tf_str}"
            )
            return
            
        symbol = args[1].upper()
        if not symbol.endswith("/USDT"):
            symbol = f"{symbol}/USDT"
            
        timeframe = args[2].lower()
        
        # Изменяем таймфрейм
        success = self.scanner.set_timeframe(symbol, timeframe)
        
        if success:
            await message.answer(f"✅ Таймфрейм для {symbol} изменен на {timeframe}")
        else:
            await message.answer(f"⚠️ Не удалось изменить таймфрейм для {symbol}. Проверьте символ и таймфрейм.")
    
    async def _cmd_scan(self, message: Message) -> None:
        """Обработчик команды /scan"""
        text = message.text
        args = text.split()
        
        if len(args) != 2:
            await message.answer("⚠️ Используйте команду в формате: /scan <символ>")
            return
            
        symbol = args[1].upper()
        if not symbol.endswith("/USDT"):
            symbol = f"{symbol}/USDT"
            
        await message.answer(f"🔍 Запуск сканирования для {symbol}...")
        
        # Запускаем сканирование
        signal = await self.scanner.scan_symbol(symbol)
        
        if signal:
            await message.answer(f"✅ Найден сигнал для {symbol}: {signal['type']} по цене {signal['price']:.4f}")
        else:
            await message.answer(f"ℹ️ Сигналов для {symbol} не найдено")
    
    async def _cmd_report(self, message: Message) -> None:
        """Обработчик команды /report - генерирует отчет о торговле"""
        try:
            # Отправляем сообщение о начале генерации отчета
            status_msg = await message.answer("📊 Генерация отчета о торговле... Запрашиваю данные с биржи")
            
            # Создаем директорию для отчетов, если её нет
            os.makedirs(REPORTS_DIR, exist_ok=True)
            
            # Путь к файлу с временной меткой последнего обновления
            timestamp_file = os.path.join(REPORTS_DIR, "last_update_timestamp.json")
            
            # Загружаем время последнего обновления или используем время запуска бота
            start_time = int(datetime.now().timestamp() * 1000) - (7 * 24 * 60 * 60 * 1000)  # По умолчанию 7 дней назад
            try:
                if os.path.exists(timestamp_file):
                    with open(timestamp_file, 'r') as f:
                        data = json.load(f)
                        start_time = data.get('last_update', start_time)
                        logger.info(f"Загружено время последнего обновления: {datetime.fromtimestamp(start_time/1000)}")
            except Exception as e:
                logger.error(f"Ошибка при загрузке времени последнего обновления: {e}")
            
            # Получаем доступные символы для запроса истории сделок
            all_trades = []
            
            try:
                # Пробуем сначала использовать TradeReporter для получения истории сделок
                await status_msg.edit_text("📊 Получаю историю сделок через TradeReporter...")
                trade_reporter = TradeReporter(self.trader.exchange.exchange)
                
                # Обновляем время последнего обновления в TradeReporter
                trade_reporter.last_update_time = start_time
                
                # Получаем новые сделки
                new_trades = await trade_reporter.fetch_new_trades()
                
                if new_trades and len(new_trades) > 0:
                    logger.info(f"Получено {len(new_trades)} сделок через TradeReporter")
                    
                    # Преобразуем сделки в формат для отчета
                    for trade in new_trades:
                        # Создаем структуру, аналогичную ccxt
                        formatted_trade = {
                            'symbol': trade.get('symbol', 'Unknown'),
                            'side': 'buy' if trade.get('side') == 'long' else 'sell',
                            'price': trade.get('price', 0),
                            'amount': trade.get('amount', 0),
                            'cost': trade.get('cost', 0),
                            'fee': {'cost': trade.get('fee', 0)},
                            'timestamp': int(trade.get('timestamp').timestamp() * 1000) if isinstance(trade.get('timestamp'), datetime) else 0,
                            'order': trade.get('trade_id', ''),
                            'info': {'pnl': trade.get('pnl', 0)}
                        }
                        all_trades.append(formatted_trade)
                else:
                    logger.info("TradeReporter не вернул новых сделок, попробуем прямой API запрос")
                    # Если TradeReporter не вернул сделок, используем прямой API запрос
                    await status_msg.edit_text("📊 Запрашиваю историю сделок напрямую с биржи...")
                    # Инициализируем пустой список для символов, _fetch_trades_via_api определит символы автоматически
                    symbols = []
                    await self._fetch_trades_via_api(all_trades, symbols, start_time, status_msg)
            except Exception as e:
                logger.error(f"Ошибка при использовании TradeReporter: {e}")
                logger.exception(e)
                
                # Если произошла ошибка, используем прямой API запрос
                await status_msg.edit_text("📊 Запрашиваю историю сделок напрямую с биржи...")
                # Инициализируем пустой список для символов
                symbols = []
                await self._fetch_trades_via_api(all_trades, symbols, start_time, status_msg)
            
            # Если нет сделок, сообщаем об этом
            if not all_trades:
                await status_msg.edit_text("ℹ️ За указанный период новых сделок не найдено.")
                return
                
            # Сортируем сделки по времени
            all_trades.sort(key=lambda x: x['timestamp'], reverse=True)
            
            # Сохраняем время последнего обновления
            latest_timestamp = max(trade['timestamp'] for trade in all_trades)
            with open(timestamp_file, 'w') as f:
                json.dump({"last_update": latest_timestamp}, f)
                logger.info(f"Сохранено время последнего обновления: {datetime.fromtimestamp(latest_timestamp/1000)}")
            
            # Создаем структурированные данные для отчета
            trade_data = []
            
            # Группируем сделки по ордерам
            orders = {}
            for trade in all_trades:
                order_id = trade.get('order')
                if order_id not in orders:
                    orders[order_id] = []
                orders[order_id].append(trade)
            
            # Обрабатываем сделки по каждому ордеру
            for order_id, trades in orders.items():
                if len(trades) == 0:
                    continue
                    
                # Определяем, лонг или шорт
                side = trades[0]['side']
                position_side = "LONG" if side == "buy" else "SHORT"
                symbol = trades[0]['symbol'].split(':')[0] if ':' in trades[0]['symbol'] else trades[0]['symbol']
                
                # Собираем все необходимые данные
                entry_price = sum(t['price'] * t['amount'] for t in trades) / sum(t['amount'] for t in trades)
                total_amount = sum(t['amount'] for t in trades)
                exit_price = sum(t.get('exit_price', t['price']) * t['amount'] for t in trades) / total_amount if total_amount > 0 else 0
                
                # Рассчитываем комиссию и PNL
                total_fee = sum(float(t['fee']['cost']) if t['fee'] and 'cost' in t['fee'] else 0 for t in trades)
                total_pnl = sum(float(t.get('info', {}).get('pnl', 0)) for t in trades)
                
                # Определяем время открытия и закрытия
                open_time = min(t['timestamp'] for t in trades)
                close_time = max(t['timestamp'] for t in trades)
                
                # Добавляем запись в отчет
                trade_data.append({
                    'Символ': symbol,
                    'Позиция': position_side,
                    'Объем': total_amount,
                    'Цена входа': entry_price,
                    'Цена выхода': exit_price,
                    'Комиссия': total_fee,
                    'PNL': total_pnl,
                    'Дата открытия': datetime.fromtimestamp(open_time / 1000).strftime('%Y-%m-%d %H:%M:%S'),
                    'Дата закрытия': datetime.fromtimestamp(close_time / 1000).strftime('%Y-%m-%d %H:%M:%S'),
                    'ID ордера': order_id
                })
            
            # Создаем DataFrame из собранных данных
            df = pd.DataFrame(trade_data)
            
            # Если данных нет, сообщаем об этом
            if len(df) == 0:
                await status_msg.edit_text("ℹ️ За указанный период не найдено завершенных сделок.")
                return
            
            # Путь к отчету
            excel_path = os.path.join(REPORTS_DIR, TRADES_EXCEL_FILE)
            
            # Обновляем статус
            await status_msg.edit_text(f"📊 Формирую отчет для {len(df)} сделок...")
            
            # Создаем Excel-отчет с красивым форматированием
            wb = Workbook()
            ws = wb.active
            ws.title = "Торговые сделки"
            
            # Записываем заголовки
            headers = list(df.columns)
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(
                    left=Side(style="thin"), 
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin")
                )
            
            # Записываем данные
            for row_idx, row in enumerate(df.values, 2):
                for col_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = Border(
                        left=Side(style="thin"), 
                        right=Side(style="thin"),
                        top=Side(style="thin"),
                        bottom=Side(style="thin")
                    )
                    
                    # Добавляем цветовое форматирование для PNL
                    if headers[col_idx-1] == "PNL":
                        if value > 0:
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        elif value < 0:
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            # Настраиваем ширину столбцов
            for col_idx, column in enumerate(ws.columns, 1):
                max_length = 0
                column_letter = get_column_letter(col_idx)
                
                for cell in column:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                
                adjusted_width = max_length + 4
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Добавляем лист со статистикой
            stats_ws = wb.create_sheet(title="Статистика")
            
            # Рассчитываем статистику
            total_trades = len(df)
            winning_trades = len(df[df["PNL"] > 0])
            losing_trades = len(df[df["PNL"] < 0])
            win_rate = (winning_trades / total_trades * 100) if total_trades > 0 else 0
            total_pnl = df["PNL"].sum()
            avg_profit = df[df["PNL"] > 0]["PNL"].mean() if winning_trades > 0 else 0
            avg_loss = df[df["PNL"] < 0]["PNL"].mean() if losing_trades > 0 else 0
            total_fee = df["Комиссия"].sum()
            
            # Статистика по символам
            symbol_stats = df.groupby("Символ").agg({
                "PNL": ["sum", "mean", "count"],
                "Комиссия": "sum"
            })
            
            # Записываем общую статистику
            stats_data = [
                ["Общая статистика", ""],
                ["Всего сделок", total_trades],
                ["Прибыльные сделки", winning_trades],
                ["Убыточные сделки", losing_trades],
                ["Процент успеха", f"{win_rate:.2f}%"],
                ["Общий PNL", f"{total_pnl:.4f} USDT"],
                ["Средняя прибыль", f"{avg_profit:.4f} USDT"],
                ["Средний убыток", f"{avg_loss:.4f} USDT"],
                ["Общая комиссия", f"{total_fee:.4f} USDT"],
                ["", ""],
                ["Статистика по символам", ""]
            ]
            
            for r_idx, row in enumerate(stats_data, 1):
                for c_idx, value in enumerate(row, 1):
                    cell = stats_ws.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 1 or r_idx == 11:
                        cell.font = Font(bold=True)
                        if c_idx == 1:
                            cell.fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
                            cell.font = Font(bold=True, color="FFFFFF")
            
            # Добавляем статистику по символам
            headers = ["Символ", "Всего сделок", "Общий PNL", "Средний PNL", "Комиссия"]
            for c_idx, header in enumerate(headers, 1):
                cell = stats_ws.cell(row=12, column=c_idx, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
            
            row_idx = 13
            for symbol, data in symbol_stats.iterrows():
                stats_ws.cell(row=row_idx, column=1, value=symbol)
                stats_ws.cell(row=row_idx, column=2, value=data[("PNL", "count")])
                stats_ws.cell(row=row_idx, column=3, value=data[("PNL", "sum")])
                stats_ws.cell(row=row_idx, column=4, value=data[("PNL", "mean")])
                stats_ws.cell(row=row_idx, column=5, value=data[("Комиссия", "sum")])
                
                # Добавляем цветовое форматирование для PNL
                pnl_cell = stats_ws.cell(row=row_idx, column=3)
                if pnl_cell.value > 0:
                    pnl_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif pnl_cell.value < 0:
                    pnl_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                
                row_idx += 1
            
            # Настраиваем ширину столбцов в листе статистики
            for col_idx, column in enumerate(stats_ws.columns, 1):
                max_length = 0
                column_letter = get_column_letter(col_idx)
                
                for cell in column:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                
                adjusted_width = max_length + 4
                stats_ws.column_dimensions[column_letter].width = adjusted_width
            
            # Сохраняем файл
            wb.save(excel_path)
            
            # Отправляем отчет
            caption = f"📊 Отчет о торговле за период с {datetime.fromtimestamp(start_time/1000).strftime('%Y-%m-%d')} по {datetime.now().strftime('%Y-%m-%d')}\n"
            caption += f"📈 Всего сделок: {total_trades}, Профит: {total_pnl:.2f} USDT, Винрейт: {win_rate:.1f}%"
            
            # Создаем объект FSInputFile вместо открытия файла напрямую
            file = FSInputFile(excel_path)
            await message.answer_document(
                document=file,
                caption=caption
            )
            
            await status_msg.delete()  # Удаляем промежуточное сообщение
            logger.info(f"Отчет о торговле успешно сгенерирован и отправлен: {excel_path}")
            
        except Exception as e:
            logger.error(f"Ошибка при генерации отчета: {e}")
            logger.exception(e)
            await message.answer(f"⚠️ Ошибка при генерации отчета: {str(e)}")
            try:
                if 'status_msg' in locals():
                    await status_msg.edit_text(f"⚠️ Ошибка при генерации отчета: {str(e)}")
            except:
                pass

    async def _fetch_trades_via_api(self, all_trades: list, symbols: list, start_time: int, status_msg: Message) -> None:
        """
        Вспомогательный метод для запроса сделок через API биржи
        
        Args:
            all_trades: Список для хранения полученных сделок
            symbols: Список символов для запроса (может быть пустым)
            start_time: Время начала запроса
            status_msg: Сообщение для обновления статуса
        """
        try:
            # Инициализируем symbols, если пустой или None
            if not symbols:
                symbols = []
                
            # Проверка существующих сделок через orders
            await status_msg.edit_text("📊 Проверяю историю завершенных ордеров...")
            
            # Получаем список завершенных ордеров для определения активных символов
            try:
                # Попытка получить историю ордеров
                closed_orders = await self.trader.exchange.exchange.fetch_closed_orders(
                    params={"instType": "swap", "marginCoin": "USDT", "limit": 50}
                )
                
                if closed_orders:
                    order_symbols = set(order['symbol'] for order in closed_orders if 'symbol' in order)
                    symbols = list(order_symbols) + symbols
                    logger.info(f"Получено {len(closed_orders)} завершенных ордеров для символов: {order_symbols}")
            except Exception as e:
                logger.error(f"Ошибка при получении закрытых ордеров: {e}")
                logger.exception(e)
        
            # Получаем список активных позиций для определения символов
            if not symbols:
                positions = await self.trader.get_active_positions()
                symbols = [pos['symbol'] for pos in positions if not pos['symbol'].endswith(':USDT')]
            
            # Если нет активных позиций, используем стандартные символы
            if not symbols:
                symbols = ["BTC/USDT", "ETH/USDT"]
            
            # Удаляем дубликаты
            symbols = list(set(symbols))
            
            await status_msg.edit_text(f"📊 Запрашиваю историю сделок для {len(symbols)} символов...")
            
            # Пытаемся сначала получить все сделки без указания символа
            try:
                all_symbol_trades = await self.trader.exchange.exchange.fetch_my_trades(
                    params={"marginCoin": "USDT", "limit": 100}
                )
                
                if all_symbol_trades:
                    all_trades.extend(all_symbol_trades)
                    logger.info(f"Получено {len(all_symbol_trades)} сделок для всех символов")
                    return
            except Exception as e:
                logger.warning(f"Ошибка при запросе всех сделок: {e}, продолжаем с отдельными символами")
            
            # Запрашиваем историю сделок для каждого символа отдельно
            for symbol in symbols:
                try:
                    symbol_trades = await self.trader.exchange.exchange.fetch_my_trades(
                        symbol=symbol, 
                        since=start_time,
                        params={"marginCoin": "USDT"}
                    )
                    
                    if symbol_trades:
                        all_trades.extend(symbol_trades)
                        logger.info(f"Получено {len(symbol_trades)} сделок для {symbol}")
                except Exception as e:
                    logger.error(f"Ошибка при запросе сделок для {symbol}: {e}")
            
            # Если не удалось получить сделки, пробуем другие методы
            if not all_trades:
                logger.warning("Не удалось получить сделки через стандартные методы, пробуем альтернативные пути")
                
                try:
                    # Прямой вызов API для получения истории сделок
                    orders_history = await self.trader.exchange.exchange.private_get_mix_order_history({
                        "instType": "swap", 
                        "marginCoin": "USDT",
                        "startTime": str(start_time),
                        "limit": "100"
                    })
                    
                    if orders_history and 'data' in orders_history:
                        logger.info(f"Получена история ордеров: {len(orders_history['data'])} ордеров")
                        # Конвертируем ордера в формат сделок
                        for order in orders_history['data']:
                            if order.get('state') == 'filled':
                                trade = {
                                    'id': order.get('orderId', ''),
                                    'order': order.get('orderId', ''),
                                    'symbol': order.get('symbol', ''),
                                    'side': order.get('side', ''),
                                    'price': float(order.get('price', 0)),
                                    'amount': float(order.get('size', 0)),
                                    'cost': float(order.get('size', 0)) * float(order.get('price', 0)),
                                    'fee': {'cost': float(order.get('fee', 0))},
                                    'timestamp': int(order.get('cTime', 0)),
                                    'info': {'pnl': float(order.get('profit', 0))}
                                }
                                all_trades.append(trade)
                except Exception as e:
                    logger.error(f"Ошибка при получении истории ордеров: {e}")
                    logger.exception(e)
        except Exception as e:
            logger.error(f"Ошибка в _fetch_trades_via_api: {e}")
            logger.exception(e)

    def register_reload_data_handler(self, callback: Callable) -> None:
        """
        Регистрирует функцию обратного вызова для перезагрузки исторических данных.
        
        Args:
            callback: Функция для перезагрузки данных
        """
        self._reload_data_callback = callback
        logger.info("Зарегистрирован обработчик перезагрузки данных")
    
    async def _cmd_reload_data(self, message: Message) -> None:
        """
        Обработчик команды /reload_data
        Перезагружает исторические данные для BTC и ETH.
        
        Args:
            message: Объект сообщения
        """
        try:
            # Проверяем, зарегистрирован ли обработчик
            if not hasattr(self, '_reload_data_callback'):
                await message.reply("⚠️ Функция перезагрузки данных не зарегистрирована")
                return
            
            # Отправляем начальное сообщение
            status_msg = await message.reply("⏳ Запуск загрузки исторических данных для BTC/USDT и ETH/USDT...")
            
            # Получаем параметры из сообщения, если они есть
            args = message.text.split()
            base_timeframe = "15m"  # По умолчанию используем 15m
            limit = 1000  # По умолчанию загружаем 1000 свечей
            
            # Если указаны параметры, используем их
            if len(args) > 1:
                base_timeframe = args[1]
            if len(args) > 2 and args[2].isdigit():
                limit = int(args[2])
                
            # Вызываем функцию перезагрузки данных с параметрами
            result = await self._reload_data_callback(base_timeframe=base_timeframe, limit=limit)
            
            # Формируем сообщение о результате
            if result["loaded"] > 0:
                details = []
                for symbol, data in result["details"].items():
                    if "error" in data:
                        details.append(f"❌ {symbol}: {data['error']}")
                    else:
                        details.append(
                            f"✅ {symbol}: загружено {data['candles']} свечей\n"
                            f"   с {data['from']} по {data['to']}"
                        )
                
                details_text = "\n\n".join(details)
                success_msg = (
                    f"✅ Успешно загружены данные для {result['loaded']} символов:\n\n"
                    f"{details_text}\n\n"
                    f"⚙️ Используемый базовый таймфрейм: {base_timeframe}\n"
                    f"⚙️ Целевой таймфрейм: 45m\n"
                    f"⚙️ Лимит свечей: {limit}"
                )
                
                await status_msg.edit_text(success_msg)
            else:
                error_msg = f"❌ Не удалось загрузить данные: {result.get('error', 'неизвестная ошибка')}"
                await status_msg.edit_text(error_msg)
                
        except Exception as e:
            logger.error(f"Ошибка при перезагрузке исторических данных: {str(e)}")
            await message.reply(f"⚠️ Ошибка при перезагрузке данных: {str(e)}")

    async def _cmd_check_indicators(self, message: Message) -> None:
        """
        Проверяет, рассчитаны ли индикаторы для предзагруженных данных BTC и ETH.
        
        Args:
            message: Объект сообщения
        """
        try:
            # Проверяем, есть ли доступ к загрузчику данных
            if not hasattr(self, '_reload_data_callback') or not hasattr(self, 'data_loader'):
                await message.reply("⚠️ Нет доступа к загрузчику данных. Сначала обновите данные командой /reload_data")
                return
            
            # Отправляем начальное сообщение
            status_msg = await message.reply("🔍 Проверяю наличие индикаторов в предзагруженных данных...")
            
            # Символы для проверки
            symbols = ["BTC/USDT", "ETH/USDT"]
            results = {}
            
            # Проверяем каждый символ
            for symbol in symbols:
                # Проверяем индикаторы в данных
                result = await self.data_loader.verify_indicators(symbol)
                results[symbol] = result
            
            # Формируем сообщение с результатами
            response = "📊 Результаты проверки индикаторов:\n\n"
            
            for symbol, result in results.items():
                if result["verified"]:
                    response += f"✅ {symbol}: Все индикаторы рассчитаны\n"
                    response += f"   Найдены: {', '.join(result['indicators_present'])}\n\n"
                else:
                    response += f"❌ {symbol}: {result['message']}\n"
                    if result["indicators_present"]:
                        response += f"   Найдены: {', '.join(result['indicators_present'])}\n"
                    if result["indicators_missing"]:
                        response += f"   Отсутствуют: {', '.join(result['indicators_missing'])}\n\n"
            
            # Отправляем результаты
            await status_msg.edit_text(response)
            
        except Exception as e:
            logger.error(f"Ошибка при проверке индикаторов: {str(e)}")
            await message.reply(f"⚠️ Ошибка при проверке индикаторов: {str(e)}") 